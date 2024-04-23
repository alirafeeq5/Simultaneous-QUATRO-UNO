import tkinter as tk
from tkinter import messagebox
import random
import openpyxl
from openpyxl.styles import PatternFill
import os

class SimultaneousQuatroUno:
    def __init__(self, root):
        self.root = root
        self.root.title("Simultaneous QUATRO-UNO")

        self.player_selection = []
        self.computer_selection = []
        self.moves = []

        self.cards = ["1", "2", "3", "4"]
        self.payoff_matrix = {
            "1": {"1": (1, 1), "2": (0, 2), "3": (0, 2), "4": (1, 1)},
            "2": {"1": (2, 0), "2": (1, 1), "3": (0, 2), "4": (0, 2)},
            "3": {"1": (2, 0), "2": (2, 0), "3": (1, 1), "4": (0, 2)},
            "4": {"1": (1, 1), "2": (2, 0), "3": (2, 0), "4": (1, 1)}
        }
        self.create_board()

    def create_board(self):
        self.title_label = tk.Label(self.root, text="Simultaneous QUATRO-UNO")
        self.title_label.grid(row=0, column=0, columnspan=5, pady=10)

        self.player_frame = tk.Frame(self.root)
        self.player_frame.grid(row=1, column=0, columnspan=5, pady=10)
        self.player_label = tk.Label(self.player_frame, text="Your Selection:")
        self.player_label.pack()
        self.player_buttons = []
        for card in self.cards:
            button = tk.Button(self.player_frame, text=card, command=lambda c=card: self.select_card(1, c))
            button.pack(side=tk.LEFT, padx=5)
            self.player_buttons.append(button)

        self.player_selection_label = tk.Label(self.root, text="")
        self.player_selection_label.grid(row=2, column=0, columnspan=5)

        self.computer_frame = tk.Frame(self.root)
        self.computer_frame.grid(row=3, column=0, columnspan=5, pady=10)
        self.computer_label = tk.Label(self.computer_frame, text="Computer's Selection:")
        self.computer_label.pack()
        self.computer_selection_label = tk.Label(self.computer_frame, text="")
        self.computer_selection_label.pack()

        self.start_button = tk.Button(self.root, text="Start Game", command=self.start_game)
        self.start_button.grid(row=4, column=0, columnspan=2, pady=10)

        self.reset_button = tk.Button(self.root, text="Reset Game", command=self.reset_game)
        self.reset_button.grid(row=4, column=3, columnspan=2, pady=10)

    def select_card(self, player, card):
        if player == 1:
            if card in self.player_selection:
                messagebox.showwarning("Warning", "You have already selected this card.")
                return
            self.player_selection.append(card)
            self.update_selection_labels()

            for button in self.player_buttons:
                if button.cget("text") == card:
                    button.config(state=tk.DISABLED)

    def update_selection_labels(self):
        if len(self.player_selection) == 4:
            for button in self.player_buttons:
                button.config(state=tk.DISABLED)
        self.computer_selection_label.config(text=", ".join(self.computer_selection))
        self.player_selection_label.config(text=", ".join(self.player_selection))

    def start_game(self):
        if len(self.player_selection) != 4:
            messagebox.showerror("Error", "You must select 4 cards.")
            return

        self.computer_selection = random.sample(self.cards, 4)
        self.computer_selection_label.config(text=", ".join(self.computer_selection))
        self.moves = list(zip(self.player_selection, self.computer_selection))
        self.play_game()

    def play_game(self):
        player_pile = list(self.player_selection)
        computer_pile = list(self.computer_selection)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payoff Matrix"

        for i, card in enumerate(self.cards):
            ws.cell(row=1, column=i + 2, value=card)
            ws.cell(row=i + 2, column=1, value=card)

        # Fill in the entire payoff matrix
        for row in range(2, len(self.cards) + 2):
            for col in range(2, len(self.cards) + 2):
                player_card = self.cards[col - 2]
                computer_card = self.cards[row - 2]
                payoff_player, payoff_computer = self.payoff_matrix[player_card][computer_card]
                ws.cell(row=row, column=col).value = f"{payoff_player},{payoff_computer}"

        def game_loop():
            nonlocal player_pile, computer_pile
            while player_pile and computer_pile:
                player_card = player_pile[0]
                computer_card = computer_pile[0]
                row = self.cards.index(player_card) + 2
                col = self.cards.index(computer_card) + 2
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                if player_card == "1" and computer_card == "4":
                    player_pile.pop(0)
                    computer_pile.pop(0)
                elif player_card == "4" and computer_card == "1":
                    player_pile.pop(0)
                    computer_pile.pop(0)
                elif player_card == computer_card:
                    player_pile.pop(0)
                    computer_pile.pop(0)
                elif player_card > computer_card:
                    computer_pile.pop(0)
                else:
                    player_pile.pop(0)

                if not player_pile and not computer_pile:
                    messagebox.showinfo("Game Over", f"DRAW")
                    break
                elif not computer_pile:
                    messagebox.showinfo("Game Over", f"Player wins")
                    break
                elif not player_pile:
                    messagebox.showinfo("Game Over", f"Computer wins")
                    break

            wb.save("payoff_matrix.xlsx")
            os.system("start payoff_matrix.xlsx")

        self.root.after(100, game_loop)

    def reset_game(self):
        self.player_selection = []
        self.computer_selection = []
        self.moves = []

        for button in self.player_buttons:
            button.config(state=tk.NORMAL)
        self.computer_selection_label.config(text="")

if __name__ == "__main__":
    root = tk.Tk()
    app = SimultaneousQuatroUno(root)
    root.mainloop()
